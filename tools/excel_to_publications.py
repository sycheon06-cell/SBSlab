#!/usr/bin/env python3
"""Convert the lab's Excel publication table into data/publications.json.

- Source: data/publications_source.xlsx
- Output: data/publications.json

Rules (as requested):
- 2 buckets only: Journal vs Proceedings
- All non-journal outputs are treated as Proceedings
- Sort: year DESC, and within the same year latest-first
- No PDF/DOI/Google Scholar links are produced (text only)
- Add (IF, Top%) in parentheses when available (typically journals)

Note: Some rows may have missing dates ('-'). In that case, the script infers
      the year by carrying forward the last known year in the sheet order.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional, Tuple

import openpyxl

# Avoid false positives inside long digit sequences (e.g., article numbers like 062007).
# This still matches patterns like "IAQVEC2024" (letter before year), "(2024)", " 2017 ", etc.
RE_YEAR = re.compile(r"(?<!\d)(19|20)\d{2}")
RE_DATE_DOT = re.compile(r"^(?P<y>(19|20)\d{2})\.(?P<m>\d{2})\.(?P<d>\d{2})$")


@dataclass
class PubRow:
    sheet_pos: int
    kind: str  # 'journal' | 'proceedings'
    year: int
    date_int: int  # YYYYMMDD, 0 if unknown
    title: str
    authors: str
    venue_text: str


def _as_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _is_missing(x: Any) -> bool:
    s = _as_str(x)
    return s == "" or s == "-" or s.lower() == "nan"


def _parse_year_from_any(*fields: Any) -> Optional[int]:
    for f in fields:
        s = _as_str(f)
        m = RE_YEAR.search(s)
        if m:
            try:
                return int(m.group(0))
            except Exception:
                pass
    return None


def _parse_date_int(date_val: Any) -> int:
    """Return YYYYMMDD int if parseable; else 0."""
    if date_val is None:
        return 0

    # Excel cell may already be a datetime
    if isinstance(date_val, datetime):
        return date_val.year * 10000 + date_val.month * 100 + date_val.day

    s = _as_str(date_val)
    if _is_missing(s):
        return 0

    m = RE_DATE_DOT.match(s)
    if not m:
        # fall back: try just year
        y = _parse_year_from_any(s)
        if y:
            return y * 10000
        return 0

    y = int(m.group("y"))
    mo = int(m.group("m"))
    d = int(m.group("d"))
    return y * 10000 + mo * 100 + d


def _fmt_if_top(if_val: Any, top_val: Any) -> str:
    """Return a formatted suffix like ' (IF 6.9, Top 4.95%)' or '' if empty."""
    if _is_missing(if_val) and _is_missing(top_val):
        return ""

    parts = []

    if not _is_missing(if_val):
        try:
            if isinstance(if_val, (int, float)):
                parts.append(f"IF {if_val:.1f}")
            else:
                parts.append(f"IF {_as_str(if_val)}")
        except Exception:
            parts.append(f"IF {_as_str(if_val)}")

    if not _is_missing(top_val):
        try:
            if isinstance(top_val, (int, float)):
                parts.append(f"Top {top_val:.2f}%")
            else:
                parts.append(f"Top {_as_str(top_val)}%")
        except Exception:
            parts.append(f"Top {_as_str(top_val)}%")

    if not parts:
        return ""

    return " (" + ", ".join(parts) + ")"


def load_excel_rows(xlsx_path: Path) -> Tuple[list[PubRow], list[PubRow]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Pick the sheet that contains '논문' in its name, else first.
    sheet_name = None
    for n in wb.sheetnames:
        if "논문" in n:
            sheet_name = n
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    journal: list[PubRow] = []
    proceedings: list[PubRow] = []

    last_year: Optional[int] = None

    # Expected columns by position (1-indexed):
    #  1 연번
    #  2 논문(Title)
    #  3 논문 구분
    #  4 발표지
    #  5 발표일자
    #  6 발표자
    #  7 (details: volume/pages)
    # 10 IF
    # 11 Top %

    for sheet_pos, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title = _as_str(row[1]) if len(row) > 1 else ""
        if title == "":
            continue

        paper_type = _as_str(row[2]) if len(row) > 2 else ""
        paper_type_l = paper_type.lower()

        # Filter out stray rows that are not journal/conference records
        if "journal" not in paper_type_l and "conference" not in paper_type_l:
            continue

        venue_main = _as_str(row[3]) if len(row) > 3 else ""
        date_val = row[4] if len(row) > 4 else None
        authors = _as_str(row[5]) if len(row) > 5 else ""
        venue_detail = _as_str(row[6]) if len(row) > 6 else ""

        if_val = row[9] if len(row) > 9 else None
        top_val = row[10] if len(row) > 10 else None

        # Year inference
        year = _parse_year_from_any(date_val, venue_main, venue_detail, title)
        if year is None:
            # Carry-forward based on the sheet order (sheet is already roughly newest->oldest)
            if last_year is not None:
                year = last_year
            else:
                # If the first record has no year (unlikely), park it at year 0.
                year = 0
        else:
            last_year = year

        date_int = _parse_date_int(date_val)

        # Venue text (match site's existing style: "JournalName volume/pages")
        venue_text = venue_main
        if venue_detail and venue_detail != "-":
            if venue_text:
                venue_text = f"{venue_text} {venue_detail}"
            else:
                venue_text = venue_detail

        # Add IF/Top% suffix if available (mostly journals)
        # We only append for journals to avoid clutter on proceedings.
        if "journal" in paper_type_l:
            venue_text = venue_text + _fmt_if_top(if_val, top_val)

        kind = "journal" if "journal" in paper_type_l else "proceedings"

        rec = PubRow(
            sheet_pos=sheet_pos,
            kind=kind,
            year=year,
            date_int=date_int,
            title=title,
            authors=authors,
            venue_text=venue_text,
        )

        if kind == "journal":
            journal.append(rec)
        else:
            proceedings.append(rec)

    return journal, proceedings


def sort_and_assign_within_year(items: list[PubRow]) -> list[dict[str, Any]]:
    # Sort: year DESC, date DESC, then sheet_pos ASC
    items_sorted = sorted(items, key=lambda x: (-x.year, -x.date_int, x.sheet_pos))

    # within_year_order: 0 is newest within the year
    output: list[dict[str, Any]] = []
    current_year = None
    within = 0

    for it in items_sorted:
        if current_year != it.year:
            current_year = it.year
            within = 0

        output.append({
            "year": it.year if it.year != 0 else None,
            "title": it.title,
            "authors": it.authors,
            "venue": it.venue_text,
            "within_year_order": within,
        })
        within += 1

    return output


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    xlsx_path = repo_root / "data" / "publications_source.xlsx"
    out_path = repo_root / "data" / "publications.json"

    journal_rows, proceedings_rows = load_excel_rows(xlsx_path)

    journal = sort_and_assign_within_year(journal_rows)
    proceedings = sort_and_assign_within_year(proceedings_rows)

    payload = {
        "generated_from": str(xlsx_path.name),
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "counts": {
            "journal": len(journal),
            "proceedings": len(proceedings),
        },
        "journal_papers": journal,
        "proceedings": proceedings,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
